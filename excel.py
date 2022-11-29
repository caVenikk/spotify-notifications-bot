import calendar
from datetime import datetime, timedelta
from openpyxl import load_workbook

from schemas import Subscriber


class ExcelHelper:
    def __init__(self, xl_path: str):
        self.xl_path = xl_path
        self.wb = load_workbook(xl_path, data_only=True)
        self.sheet = self.wb['Лист1']

    def get_subscribers(self, max: int) -> list[Subscriber]:
        subs = []
        for cellObj in self.sheet['A2':f'H{max}']:
            values = []
            for cell in cellObj:
                if cell.value != None:
                    values.append(cell.value)
            if values:
                sub = Subscriber(*values)
                subs.append(sub)
            else:
                break
        return subs

    @staticmethod
    def get_outdated_subscribers(subs: list[Subscriber]) -> list[Subscriber]:
        now = datetime.today()
        res = []
        for sub in subs:
            date = sub.date
            days_in_month = calendar.monthrange(date.year, date.month)[1]
            if now.date() == date + timedelta(days=days_in_month-1):
                res.append(sub)
        return res

    def update(self, xl_path=None):
        if not xl_path:
            xl_path = self.xl_path
        self.xl_path = xl_path
        self.wb = load_workbook(xl_path, data_only=True)
        self.sheet = self.wb['Лист1']

    def get_earnings(self) -> str:
        return f"{self.sheet['N2'].value} rub"

    def get_clean_earnings(self) -> str:
        return f"{self.sheet['O2'].value} rub"
